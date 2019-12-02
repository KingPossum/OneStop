import sys
import xlsxwriter
import pyexcel
import random
import openpyxl
from tkinter import *
from tkinter import filedialog, messagebox
from openpyxl.styles import Alignment
import tkinter.ttk as ttk
py3 = 1

def OnClosing():
    if messagebox.askokcancel("WARNING", "Are you sure you want to quit?\nUnsaved data will be lost."):
        root.destroy()
def vp_start_gui():
    '''Starting point when module is the main routine.'''
    global val, w, root
    root = Tk()
    root.iconbitmap("icon.ico")
    top = OneStop (root)
    root.protocol("WM_DELETE_WINDOW", OnClosing)
    root.mainloop()

w = None
def create_OneStop(root, *args, **kwargs):
    '''Starting point when module is imported by another program.'''
    global w, w_win, rt
    rt = root
    w = Toplevel (root)
    top = OneStop (w)

    return (w, top)

def destroy_OneStop():
    global w
    w.destroy()
    w = None


class OneStop:
    def __init__(self, top=None):
        '''This class configures and populates the toplevel window.
           top is the toplevel containing window.'''
        _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
        _fgcolor = '#000000'  # X11 color: 'black'
        _compcolor = '#d9d9d9' # X11 color: 'gray85'
        _ana1color = '#d9d9d9' # X11 color: 'gray85' 
        _ana2color = '#d9d9d9' # X11 color: 'gray85' 
        self.style = ttk.Style()
        if sys.platform == "win32":
            self.style.theme_use('winnative')
        self.style.configure('.',background=_bgcolor)
        self.style.configure('.',foreground=_fgcolor)
        self.style.map('.',background=
            [('selected', _compcolor), ('active',_ana2color)])

        top.geometry("534x309+637+323")
        top.title("OneStop")
        top.configure(background="#d9d9d9")
        top.configure(highlightbackground="#d9d9d9")
        top.configure(highlightcolor="black")
        top.bind("<Delete>", self.DelEntry)
        top.bind("<Insert>", self.SaveEntry)
        top.bind("<Home>", self.RollRandom)
        #top.bind("<End>", self.AddEntry)
        
        self.ItemList = []

        def onSelect(event):
            
            self.Description.delete('1.0',END)
            self.BasePrice.delete('0',END)
            self.HighPrice.delete('0',END)
            self.LowPrice.delete('0',END)
            self.RandPrice.delete('0',END)
            current = self.ShopInventory.get(self.ShopInventory.curselection())
            try:
                result = next(sublist for sublist in self.ItemList if sublist[0] == current) #the sublist pulled out of main self.ItemList
            except StopIteration as e:
                self.ItemList = self.ItemList[0]
                result = next(sublist for sublist in self.ItemList if sublist[0] == current)
    
                
            
            if self.Description.get('1.0', END) == result[1]:
                pass
            else:
                self.Description.insert(END,result[1])
                self.BasePrice.insert(END,result[2])
                self.HighPrice.insert(END,result[3])
                self.LowPrice.insert(END,result[4])
                self.RandPrice.insert(END,result[5])
                
                
        self.ShopInventory = ScrolledListBox(top)
        self.ShopInventory.place(relx=0.0, rely=0.02, relheight=0.96  
                , relwidth=0.4)
        self.ShopInventory.configure(background="white")
        self.ShopInventory.configure(disabledforeground="#a3a3a3")
        self.ShopInventory.configure(font="TkFixedFont")
        self.ShopInventory.configure(foreground="black")
        self.ShopInventory.configure(highlightbackground="#d9d9d9")
        self.ShopInventory.configure(highlightcolor="#d9d9d9")
        self.ShopInventory.configure(selectbackground="#c4c4c4")
        self.ShopInventory.configure(selectforeground="black")
        self.ShopInventory.configure(width=10)
        self.ShopInventory.bind("<<ListboxSelect>>", onSelect)
        
        self.ImportButton = Button(top, command=self.ImportFile)
        self.ImportButton.place(relx=0.41, rely=0.91, height=24, width=87)
        self.ImportButton.configure(activebackground="#d9d9d9")
        self.ImportButton.configure(activeforeground="#000000")
        self.ImportButton.configure(background="#d9d9d9")
        self.ImportButton.configure(disabledforeground="#a3a3a3")
        self.ImportButton.configure(foreground="#000000")
        self.ImportButton.configure(highlightbackground="#d9d9d9")
        self.ImportButton.configure(highlightcolor="black")
        self.ImportButton.configure(pady="0")
        self.ImportButton.configure(text='''Open''')

        self.ExportButton = Button(top, command = self.ExportFile)
        self.ExportButton.place(relx=0.82, rely=0.91, height=24, width=87)
        self.ExportButton.configure(activebackground="#d9d9d9")
        self.ExportButton.configure(activeforeground="#000000")
        self.ExportButton.configure(background="#d9d9d9")
        self.ExportButton.configure(disabledforeground="#a3a3a3")
        self.ExportButton.configure(foreground="#000000")
        self.ExportButton.configure(highlightbackground="#d9d9d9")
        self.ExportButton.configure(highlightcolor="black")
        self.ExportButton.configure(pady="0")
        self.ExportButton.configure(text='''Save As''')

        self.SaveEntry = Button(top, command = self.SaveEntry)
        self.SaveEntry.place(relx=0.62, rely=0.78, height=24, width=87)
        self.SaveEntry.configure(activebackground="#d9d9d9")
        self.SaveEntry.configure(activeforeground="#000000")
        self.SaveEntry.configure(background="#d9d9d9")
        self.SaveEntry.configure(disabledforeground="#a3a3a3")
        self.SaveEntry.configure(foreground="#000000")
        self.SaveEntry.configure(highlightbackground="#d9d9d9")
        self.SaveEntry.configure(highlightcolor="black")
        self.SaveEntry.configure(pady="0")
        self.SaveEntry.configure(text='''Save Entry''')

        self.BaseN = Label(top)
        self.BaseN.place(relx=0.41, rely=0.58, height=31, width=64)
        self.BaseN.configure(activebackground="#f9f9f9")
        self.BaseN.configure(activeforeground="black")
        self.BaseN.configure(background="#d9d9d9")
        self.BaseN.configure(disabledforeground="#a3a3a3")
        self.BaseN.configure(foreground="#000000")
        self.BaseN.configure(highlightbackground="#d9d9d9")
        self.BaseN.configure(highlightcolor="black")
        self.BaseN.configure(text='''Base''')
        self.BaseN.configure(width=64)

        self.HighN = Label(top)
        self.HighN.place(relx=0.56, rely=0.58, height=31, width=64)
        self.HighN.configure(activebackground="#f9f9f9")
        self.HighN.configure(activeforeground="black")
        self.HighN.configure(background="#d9d9d9")
        self.HighN.configure(disabledforeground="#a3a3a3")
        self.HighN.configure(foreground="#000000")
        self.HighN.configure(highlightbackground="#d9d9d9")
        self.HighN.configure(highlightcolor="black")
        self.HighN.configure(text='''High''')
        self.HighN.configure(width=64)

        self.LowN = Label(top)
        self.LowN.place(relx=0.71, rely=0.58, height=31, width=64)
        self.LowN.configure(activebackground="#f9f9f9")
        self.LowN.configure(activeforeground="black")
        self.LowN.configure(background="#d9d9d9")
        self.LowN.configure(disabledforeground="#a3a3a3")
        self.LowN.configure(foreground="#000000")
        self.LowN.configure(highlightbackground="#d9d9d9")
        self.LowN.configure(highlightcolor="black")
        self.LowN.configure(text='''Low''')
        self.LowN.configure(width=64)

        self.RandN = Label(top)
        self.RandN.place(relx=0.86, rely=0.58, height=31, width=64)
        self.RandN.configure(activebackground="#f9f9f9")
        self.RandN.configure(activeforeground="black")
        self.RandN.configure(background="#d9d9d9")
        self.RandN.configure(disabledforeground="#a3a3a3")
        self.RandN.configure(foreground="#000000")
        self.RandN.configure(highlightbackground="#d9d9d9")
        self.RandN.configure(highlightcolor="black")
        self.RandN.configure(text='''Rand''')
        self.RandN.configure(width=64)

        self.ReRoll = Button(top, command = self.RollRandom)
        self.ReRoll.place(relx=0.82, rely=0.78, height=24, width=87)
        self.ReRoll.configure(activebackground="#d9d9d9")
        self.ReRoll.configure(activeforeground="#000000")
        self.ReRoll.configure(background="#d9d9d9")
        self.ReRoll.configure(disabledforeground="#a3a3a3")
        self.ReRoll.configure(foreground="#000000")
        self.ReRoll.configure(highlightbackground="#d9d9d9")
        self.ReRoll.configure(highlightcolor="black")
        self.ReRoll.configure(pady="0")
        self.ReRoll.configure(text='''Roll Random''')

        self.AddEntry = Button(top, command = self.AddEntry)
        self.AddEntry.place(relx=0.41, rely=0.78, height=24, width=87)
        self.AddEntry.configure(activebackground="#d9d9d9")
        self.AddEntry.configure(activeforeground="#000000")
        self.AddEntry.configure(background="#d9d9d9")
        self.AddEntry.configure(disabledforeground="#a3a3a3")
        self.AddEntry.configure(foreground="#000000")
        self.AddEntry.configure(highlightbackground="#d9d9d9")
        self.AddEntry.configure(highlightcolor="black")
        self.AddEntry.configure(pady="0")
        self.AddEntry.configure(text='''Add Entry''')

        self.DeleteEntry = Button(top, command = self.DelEntry)
        self.DeleteEntry.place(relx=0.62, rely=0.91, height=24, width=87)
        self.DeleteEntry.configure(activebackground="#d9d9d9")
        self.DeleteEntry.configure(activeforeground="#000000")
        self.DeleteEntry.configure(background="#d9d9d9")
        self.DeleteEntry.configure(disabledforeground="#a3a3a3")
        self.DeleteEntry.configure(foreground="#000000")
        self.DeleteEntry.configure(highlightbackground="#d9d9d9")
        self.DeleteEntry.configure(highlightcolor="black")
        self.DeleteEntry.configure(pady="0")
        self.DeleteEntry.configure(text='''Del Entry''')

        self.menubar = Menu(top,font="TkMenuFont",bg=_bgcolor,fg=_fgcolor)
        top.configure(menu = self.menubar)



        self.BasePrice = Entry(top)
        self.BasePrice.place(relx=0.41, rely=0.49, relheight=0.1, relwidth=0.12)
        self.BasePrice.configure(background="#c0c0c0")
        self.BasePrice.configure(disabledforeground="#a3a3a3")
        self.BasePrice.configure(font="TkFixedFont")
        self.BasePrice.configure(foreground="#000000")
        self.BasePrice.configure(insertbackground="black")
        self.BasePrice.configure(textvariable="BPVAR")
        self.BasePrice.configure(width=64)

        self.HighPrice = Entry(top)
        self.HighPrice.place(relx=0.56, rely=0.49, relheight=0.1, relwidth=0.12)
        self.HighPrice.configure(background="#c0c0c0")
        self.HighPrice.configure(disabledforeground="#a3a3a3")
        self.HighPrice.configure(font="TkFixedFont")
        self.HighPrice.configure(foreground="#000000")
        self.HighPrice.configure(highlightbackground="#d9d9d9")
        self.HighPrice.configure(highlightcolor="black")
        self.HighPrice.configure(insertbackground="black")
        self.HighPrice.configure(selectbackground="#c4c4c4")
        self.HighPrice.configure(selectforeground="black")
        self.HighPrice.configure(textvariable="HVAR")

        self.LowPrice = Entry(top)
        self.LowPrice.place(relx=0.71, rely=0.49, relheight=0.1, relwidth=0.12)
        self.LowPrice.configure(background="#c0c0c0")
        self.LowPrice.configure(disabledforeground="#a3a3a3")
        self.LowPrice.configure(font="TkFixedFont")
        self.LowPrice.configure(foreground="#000000")
        self.LowPrice.configure(highlightbackground="#d9d9d9")
        self.LowPrice.configure(highlightcolor="black")
        self.LowPrice.configure(insertbackground="black")
        self.LowPrice.configure(selectbackground="#c4c4c4")
        self.LowPrice.configure(selectforeground="black")
        self.LowPrice.configure(textvariable="LVAR")

        self.RandPrice = Entry(top)
        self.RandPrice.place(relx=0.86, rely=0.49, relheight=0.1, relwidth=0.12)
        self.RandPrice.configure(background="#c0c0c0")
        self.RandPrice.configure(disabledforeground="#a3a3a3")
        self.RandPrice.configure(font="TkFixedFont")
        self.RandPrice.configure(foreground="#000000")
        self.RandPrice.configure(highlightbackground="#d9d9d9")
        self.RandPrice.configure(highlightcolor="black")
        self.RandPrice.configure(insertbackground="black")
        self.RandPrice.configure(selectbackground="#c4c4c4")
        self.RandPrice.configure(selectforeground="black")
        self.RandPrice.configure(textvariable="RVAR")
        self.RandPrice.configure(width=64)

        self.Description = Text(top)
        self.Description.place(relx=0.41, rely=0.03, relheight=0.43
                , relwidth=0.57)
        self.Description.configure(background="#c0c0c0")
        self.Description.configure(font="TkTextFont")
        self.Description.configure(foreground="black")
        self.Description.configure(highlightbackground="#d9d9d9")
        self.Description.configure(highlightcolor="black")
        self.Description.configure(insertbackground="black")
        self.Description.configure(selectbackground="#c4c4c4")
        self.Description.configure(selectforeground="black")
        self.Description.configure(width=304)
        self.Description.configure(wrap=WORD)
        
    def RollRandom(self, *args):
        desc = self.Description.get('1.0','end-1c')
        name = desc.split(':')
        high = self.HighPrice.get()
        low = self.LowPrice.get()
        
        if high == '' or low == '':
            for index, sublist in enumerate(self.ItemList):
                if sublist[0] == name[0]: 
                    sublist[5] = random.randint(int(sublist[4]),int(sublist[3]))
                    self.RandPrice.delete('0',END)
                    self.RandPrice.insert(END,sublist[5])
        else:
            randNum = random.randint(int(low),int(high))
            self.RandPrice.delete('0',END)
            self.RandPrice.insert(END,randNum)
    def AddEntry(self):
        self.Description.delete('1.0',END)
        self.BasePrice.delete('0',END)
        self.HighPrice.delete('0',END)
        self.LowPrice.delete('0',END)
        self.RandPrice.delete('0',END)
        
    
    def DelEntry(self, *args):
        desc = self.Description.get('1.0','end-1c')
        name = desc.split(':')
        for index, sublist in enumerate(self.ItemList):
            if sublist[0] == name[0]:
                del self.ItemList[index]
                self.ShopInventory.delete(0, END)
                names = [item[0] for item in self.ItemList]
                for aName in names:
                    self.ShopInventory.insert(END, aName)
                    
        self.Description.delete('1.0',END)
        self.BasePrice.delete('0',END)
        self.HighPrice.delete('0',END)
        self.LowPrice.delete('0',END)
        self.RandPrice.delete('0',END)
                    
    def SaveEntry(self, *args):
        desc = self.Description.get('1.0','end-1c')       
        name = desc.split(":")
        for index, sublist in enumerate(self.ItemList): #CHECKS IF DUPLICATE SUBLIST IS IN LIST
            if sublist[0] == name[0]:
                del self.ItemList[index]
                self.ShopInventory.delete(0, END)
                names = [item[0] for item in self.ItemList]
                for aName in names:
                    self.ShopInventory.insert(END, aName)
          #######################################################THIS IS WHERE I WAS
        base = self.BasePrice.get()
        high = self.HighPrice.get()
        low = self.LowPrice.get()
        rand = self.RandPrice.get()
        item = [name[0],desc,base,high,low,rand]
        self.ItemList.append(item)      
        self.Description.delete('1.0', END)
        self.BasePrice.delete('0', END)
        self.HighPrice.delete('0', END)
        self.LowPrice.delete('0', END)
        self.RandPrice.delete('0', END)
        self.ShopInventory.insert(END, str(name[0]))

    def ExportFile(self):
        f = filedialog.asksaveasfilename(title = "Save As",filetypes=(("Spreadsheet", "*.xlsx"),("All Files","*.*")))
        if f == '':
            return
        else:

            CheckString = ".xlsx"
            if CheckString in f:
                workbook = xlsxwriter.Workbook(f)
            else:
                workbook = xlsxwriter.Workbook(f + ".xlsx")

            bold = workbook.add_format({'bold': True})
            text_format = workbook.add_format({'text_wrap': True})
            worksheet = workbook.add_worksheet()
            worksheet.set_column('B:B',38)
            worksheet.set_column('A:A',15)
            worksheet.write(0,0, "Name", bold)
            worksheet.write(0,1, "Description", bold)
            worksheet.write(0,2, "Base $", bold)
            worksheet.write(0,3, "High $", bold)
            worksheet.write(0,4, "Low $", bold)
            worksheet.write(0,5, "Rand $", bold)
            row = 1
            col = 0
            for llist in self.ItemList:
                for item in llist:
                    descCheck = ":" in item
                    if descCheck:
                        print(descCheck)
                        item = item.split(":",1)[-1]
                        
                    worksheet.write(row,col, item, text_format)
                    col +=1
                row +=1
                col = 0
            workbook.close()

            if CheckString in f:
                messagebox.showinfo("Export Successful", f.rsplit("/",1)[-1] + " created")
            else:
                messagebox.showinfo("Export Successful", f.rsplit("/",1)[-1] + ".xlsx created")
            
    def ImportFile(self):

        def _itersplit(fin, splitters):
            current = []
            for item in fin:
                if item in splitters:
                    yield current
                    current = []
                else:
                    current.append(item)
            yield current

        def magicsplit(fin, *splitters):
            return [subl for subl in _itersplit(fin, splitters) if subl]

        f = filedialog.askopenfilename(filetypes =(("Spreadsheet","*.xlsx"),("All Files","*.*")))
        if f == "":
            return
        else:
            try:
                wb = openpyxl.load_workbook(f, read_only = True)
                ws = wb.active
                pre = []
                for row in ws.iter_rows(row_offset=1):
                    if len(pre) != 0:
                        pre.append(None)
                    for cell in row:
                        value = cell.value
                        
                        if value is None:
                            print("in none!")
                            pass
                        else:
                            pre.append(value)

                del pre[-1]
                it = magicsplit(pre, None)
                print(str(it))
                for alist in it:
                    alist[1] = alist[0]+":"+alist[1]
                    del self.ItemList[:]
                    self.ItemList.append(it)
                    self.ShopInventory.delete(0, END)
                    self.Description.delete('1.0', END)
                    self.BasePrice.delete('0', END)
                    self.HighPrice.delete('0', END)
                    self.LowPrice.delete('0', END)
                    self.RandPrice.delete('0', END)
                    names = [item[0] for item in self.ItemList[0]]
                    for aName in names:
                        self.ShopInventory.insert(END, aName)
            except Exception as e:
                messagebox.showinfo("Import Failed","Spreadsheet may be damaged or corrupted. Merged cells can also raise exceptions")
                print(e)
                pass

    @staticmethod
    def popup1(event):
        Popupmenu1 = Menu(root, tearoff=0)
        Popupmenu1.configure(activebackground="#f9f9f9")
        Popupmenu1.configure(activeborderwidth="1")
        Popupmenu1.configure(activeforeground="black")
        Popupmenu1.configure(background="#d9d9d9")
        Popupmenu1.configure(borderwidth="1")
        Popupmenu1.configure(disabledforeground="#a3a3a3")
        Popupmenu1.configure(font="{Segoe UI} 9")
        Popupmenu1.configure(foreground="black")
        Popupmenu1.post(event.x_root, event.y_root)





# The following code is added to facilitate the Scrolled widgets you specified.
class AutoScroll(object):
    '''Configure the scrollbars for a widget.'''

    def __init__(self, master):
        #  Rozen. Added the try-except clauses so that this class
        #  could be used for scrolled entry widget for which vertical
        #  scrolling is not supported. 5/7/14.
        try:
            vsb = ttk.Scrollbar(master, orient='vertical', command=self.yview)
        except:
            pass
        hsb = ttk.Scrollbar(master, orient='horizontal', command=self.xview)

        #self.configure(yscrollcommand=_autoscroll(vsb),
        #    xscrollcommand=_autoscroll(hsb))
        try:
            self.configure(yscrollcommand=self._autoscroll(vsb))
        except:
            pass
        self.configure(xscrollcommand=self._autoscroll(hsb))

        self.grid(column=0, row=0, sticky='nsew')
        try:
            vsb.grid(column=1, row=0, sticky='ns')
        except:
            pass
        hsb.grid(column=0, row=1, sticky='ew')

        master.grid_columnconfigure(0, weight=1)
        master.grid_rowconfigure(0, weight=1)

        # Copy geometry methods of master  (taken from ScrolledText.py)
        if py3:
            methods = Pack.__dict__.keys() | Grid.__dict__.keys() \
                  | Place.__dict__.keys()
        else:
            methods = Pack.__dict__.keys() + Grid.__dict__.keys() \
                  + Place.__dict__.keys()

        for meth in methods:
            if meth[0] != '_' and meth not in ('config', 'configure'):
                setattr(self, meth, getattr(master, meth))

    @staticmethod
    def _autoscroll(sbar):
        '''Hide and show scrollbar as needed.'''
        def wrapped(first, last):
            first, last = float(first), float(last)
            if first <= 0 and last >= 1:
                sbar.grid_remove()
            else:
                sbar.grid()
            sbar.set(first, last)
        return wrapped

    def __str__(self):
        return str(self.master)

def _create_container(func):
    '''Creates a ttk Frame with a given master, and use this new frame to
    place the scrollbars and the widget.'''
    def wrapped(cls, master, **kw):
        container = ttk.Frame(master)
        return func(cls, container, **kw)
    return wrapped

class ScrolledListBox(AutoScroll, Listbox):
    '''A standard Tkinter Text widget with scrollbars that will
    automatically show/hide as needed.'''
    @_create_container
    def __init__(self, master, **kw):
        Listbox.__init__(self, master, **kw)
        AutoScroll.__init__(self, master)

if __name__ == '__main__':
    vp_start_gui()


